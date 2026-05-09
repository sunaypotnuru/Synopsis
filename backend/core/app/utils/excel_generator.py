"""
Excel Report Generator for NetraAI MCP Analytics
Generates comprehensive Excel reports with multiple sheets, charts, and formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List
import io
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Professional Excel report generator for MCP analytics and audit logs.

    Features:
    - Multiple sheets (Dashboard, Usage, Success, Latency, etc.)
    - Charts and graphs
    - Conditional formatting
    - Professional styling
    - Pivot-ready data
    """

    def __init__(self):
        self.wb = Workbook()
        self.header_fill = PatternFill(
            start_color="0D9488", end_color="0D9488", fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="0D9488")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_analytics_report(self, analytics_data: Dict) -> bytes:
        """
        Generate comprehensive analytics Excel report.

        Parameters:
        - analytics_data: Dictionary containing all analytics data

        Returns:
        - Excel file as bytes
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create sheets
        self._create_dashboard_sheet(analytics_data)

        if "usage_trends" in analytics_data:
            self._create_usage_trends_sheet(analytics_data["usage_trends"])

        if "success_rates" in analytics_data:
            self._create_success_rates_sheet(analytics_data["success_rates"])

        if "latency_distribution" in analytics_data:
            self._create_latency_sheet(analytics_data["latency_distribution"])

        if "geographic_distribution" in analytics_data:
            self._create_geographic_sheet(analytics_data["geographic_distribution"])

        if "error_breakdown" in analytics_data:
            self._create_error_sheet(analytics_data["error_breakdown"])

        # Save to bytes
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_audit_log_report(self, logs: List[Dict]) -> bytes:
        """
        Generate audit log Excel report.

        Parameters:
        - logs: List of audit log entries

        Returns:
        - Excel file as bytes
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create audit logs sheet
        self._create_audit_logs_sheet(logs)

        # Create summary sheet
        self._create_audit_summary_sheet(logs)

        # Save to bytes
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_dashboard_sheet(self, analytics_data: Dict):
        """Create dashboard summary sheet."""
        ws = self.wb.create_sheet("Dashboard", 0)

        # Title
        ws["A1"] = "NetraAI MCP Analytics Dashboard"
        ws["A1"].font = Font(bold=True, size=16, color="0D9488")
        ws.merge_cells("A1:D1")

        # Generated date
        ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws["A2"].font = Font(italic=True, size=10)
        ws.merge_cells("A2:D2")

        # Key Metrics
        row = 4
        ws[f"A{row}"] = "Key Performance Indicators"
        ws[f"A{row}"].font = self.title_font
        row += 1

        # Headers
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].fill = self.header_fill
        ws[f"B{row}"].fill = self.header_fill
        ws[f"A{row}"].font = self.header_font
        ws[f"B{row}"].font = self.header_font
        row += 1

        # Metrics
        metrics = []

        if "usage_trends" in analytics_data:
            usage = analytics_data["usage_trends"]
            metrics.append(
                ("Total Invocations (24h)", f"{usage.get('total_invocations', 0):,}")
            )
            metrics.append(("Peak Hour", usage.get("peak_hour", "N/A")))
            metrics.append(("Avg per Hour", f"{usage.get('avg_per_hour', 0):.1f}"))

        if "success_rates" in analytics_data:
            success = analytics_data["success_rates"]
            metrics.append(
                (
                    "Overall Success Rate",
                    f"{(success.get('overall_success_rate', 0) * 100):.1f}%",
                )
            )
            metrics.append(("Total Calls", f"{success.get('total_calls', 0):,}"))

        if "latency_distribution" in analytics_data:
            latency = analytics_data["latency_distribution"]
            percentiles = latency.get("percentiles", {})
            metrics.append(("Median Latency (P50)", f"{percentiles.get('p50', 0)}ms"))
            metrics.append(("P95 Latency", f"{percentiles.get('p95', 0)}ms"))

        if "error_breakdown" in analytics_data:
            errors = analytics_data["error_breakdown"]
            metrics.append(
                ("Error Rate", f"{(errors.get('error_rate', 0) * 100):.2f}%")
            )
            metrics.append(("Total Errors", f"{errors.get('total_errors', 0):,}"))

        for metric, value in metrics:
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value
            ws[f"A{row}"].border = self.border
            ws[f"B{row}"].border = self.border
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_usage_trends_sheet(self, usage_data: Dict):
        """Create usage trends sheet with chart."""
        ws = self.wb.create_sheet("Usage Trends")

        # Title
        ws["A1"] = "Tool Usage Trends (24 Hours)"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:G1")

        # Headers
        headers = [
            "Hour",
            "Total",
            "Anemia",
            "Cataract",
            "DR",
            "Mental Health",
            "Parkinsons",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        data = usage_data.get("data", [])
        for row_idx, hour_data in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=hour_data.get("hour", ""))
            ws.cell(row=row_idx, column=2, value=hour_data.get("total_invocations", 0))
            ws.cell(row=row_idx, column=3, value=hour_data.get("anemia", 0))
            ws.cell(row=row_idx, column=4, value=hour_data.get("cataract", 0))
            ws.cell(row=row_idx, column=5, value=hour_data.get("dr", 0))
            ws.cell(row=row_idx, column=6, value=hour_data.get("mental_health", 0))
            ws.cell(row=row_idx, column=7, value=hour_data.get("parkinsons", 0))

            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = self.border

        # Create line chart
        chart = LineChart()
        chart.title = "Usage Trends Over 24 Hours"
        chart.style = 10
        chart.y_axis.title = "Invocations"
        chart.x_axis.title = "Hour"

        # Add data to chart
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(data), max_col=7)
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "A30")

        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_success_rates_sheet(self, success_data: Dict):
        """Create success rates sheet with chart."""
        ws = self.wb.create_sheet("Success Rates")

        # Title
        ws["A1"] = "Success Rates by Tool"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:F1")

        # Headers
        headers = [
            "Tool",
            "Category",
            "Success Rate",
            "Total Calls",
            "Successful",
            "Failed",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        tools = success_data.get("tools", [])
        for row_idx, tool in enumerate(tools, 4):
            ws.cell(
                row=row_idx,
                column=1,
                value=tool.get("tool", "").replace("_tool", "").replace("_", " "),
            )
            ws.cell(row=row_idx, column=2, value=tool.get("category", ""))
            ws.cell(row=row_idx, column=3, value=tool.get("success_rate", 0))
            ws.cell(row=row_idx, column=4, value=tool.get("total_calls", 0))
            ws.cell(row=row_idx, column=5, value=tool.get("successful_calls", 0))
            ws.cell(row=row_idx, column=6, value=tool.get("failed_calls", 0))

            # Format success rate as percentage
            ws.cell(row=row_idx, column=3).number_format = "0.0%"

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = self.border

        # Create bar chart
        chart = BarChart()
        chart.title = "Success Rates by Tool"
        chart.style = 10
        chart.y_axis.title = "Success Rate"
        chart.x_axis.title = "Tool"

        # Add data to chart
        data_ref = Reference(ws, min_col=3, min_row=3, max_row=3 + len(tools))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(tools))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "A30")

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_latency_sheet(self, latency_data: Dict):
        """Create latency distribution sheet."""
        ws = self.wb.create_sheet("Latency Distribution")

        # Title
        ws["A1"] = "Latency Distribution"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Percentiles section
        ws["A3"] = "Percentiles"
        ws["A3"].font = Font(bold=True, size=12)

        percentiles = latency_data.get("percentiles", {})
        perc_data = [
            ("P50 (Median)", f"{percentiles.get('p50', 0)}ms"),
            ("P75", f"{percentiles.get('p75', 0)}ms"),
            ("P90", f"{percentiles.get('p90', 0)}ms"),
            ("P95", f"{percentiles.get('p95', 0)}ms"),
            ("P99", f"{percentiles.get('p99', 0)}ms"),
        ]

        row = 4
        for label, value in perc_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].border = self.border
            ws[f"B{row}"].border = self.border
            row += 1

        # Distribution section
        ws["A10"] = "Distribution Buckets"
        ws["A10"].font = Font(bold=True, size=12)

        # Headers
        headers = ["Latency Range", "Request Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        buckets = latency_data.get("buckets", [])
        for row_idx, bucket in enumerate(buckets, 12):
            ws.cell(row=row_idx, column=1, value=bucket.get("range", ""))
            ws.cell(row=row_idx, column=2, value=bucket.get("count", 0))
            ws.cell(row=row_idx, column=3, value=bucket.get("percentage", 0) / 100)

            # Format percentage
            ws.cell(row=row_idx, column=3).number_format = "0.0%"

            # Apply borders
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = self.border

        # Create bar chart
        chart = BarChart()
        chart.title = "Latency Distribution"
        chart.style = 10
        chart.y_axis.title = "Request Count"
        chart.x_axis.title = "Latency Range"

        data_ref = Reference(ws, min_col=2, min_row=11, max_row=11 + len(buckets))
        cats = Reference(ws, min_col=1, min_row=12, max_row=11 + len(buckets))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "E3")

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15

    def _create_geographic_sheet(self, geo_data: Dict):
        """Create geographic distribution sheet."""
        ws = self.wb.create_sheet("Geographic Distribution")

        # Title
        ws["A1"] = "Geographic Distribution"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:D1")

        # Headers
        headers = ["Region", "Requests", "Percentage", "Avg Latency (ms)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        regions = geo_data.get("regions", [])
        for row_idx, region in enumerate(regions, 4):
            ws.cell(row=row_idx, column=1, value=region.get("region", ""))
            ws.cell(row=row_idx, column=2, value=region.get("requests", 0))
            ws.cell(row=row_idx, column=3, value=region.get("percentage", 0) / 100)
            ws.cell(row=row_idx, column=4, value=region.get("avg_latency_ms", 0))

            # Format percentage
            ws.cell(row=row_idx, column=3).number_format = "0.0%"

            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = self.border

        # Create pie chart
        chart = PieChart()
        chart.title = "Requests by Region"
        chart.style = 10

        data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(regions))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(regions))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "F3")

        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_error_sheet(self, error_data: Dict):
        """Create error breakdown sheet."""
        ws = self.wb.create_sheet("Error Breakdown")

        # Title
        ws["A1"] = "Error Breakdown"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:D1")

        # Headers
        headers = ["Error Type", "Count", "Percentage", "Severity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        error_types = error_data.get("error_types", [])
        for row_idx, error in enumerate(error_types, 4):
            ws.cell(row=row_idx, column=1, value=error.get("type", ""))
            ws.cell(row=row_idx, column=2, value=error.get("count", 0))
            ws.cell(row=row_idx, column=3, value=error.get("percentage", 0) / 100)
            ws.cell(row=row_idx, column=4, value=error.get("severity", ""))

            # Format percentage
            ws.cell(row=row_idx, column=3).number_format = "0.0%"

            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = self.border

        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_audit_logs_sheet(self, logs: List[Dict]):
        """Create audit logs sheet."""
        ws = self.wb.create_sheet("Audit Logs", 0)

        # Title
        ws["A1"] = "Audit Log Entries"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:F1")

        # Headers
        headers = [
            "Timestamp",
            "Tool Name",
            "Status",
            "Patient ID",
            "Latency (ms)",
            "Event Type",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Data
        for row_idx, log in enumerate(logs, 4):
            timestamp = log.get("timestamp", "")
            if timestamp:
                try:
                    dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                    timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
                except (ValueError, AttributeError):
                    pass

            ws.cell(row=row_idx, column=1, value=timestamp)
            ws.cell(row=row_idx, column=2, value=log.get("tool_name", ""))
            ws.cell(row=row_idx, column=3, value=log.get("status", ""))
            ws.cell(row=row_idx, column=4, value=log.get("patient_id", ""))
            ws.cell(row=row_idx, column=5, value=log.get("latency_ms", 0))
            ws.cell(row=row_idx, column=6, value=log.get("event_type", ""))

            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = self.border

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 18

        # Enable auto-filter
        ws.auto_filter.ref = f"A3:F{3 + len(logs)}"

    def _create_audit_summary_sheet(self, logs: List[Dict]):
        """Create audit summary sheet."""
        ws = self.wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Audit Log Summary"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:B1")

        # Calculate summary stats
        total_logs = len(logs)
        success_logs = len([log for log in logs if log.get("status") == "SUCCESS"])
        error_logs = total_logs - success_logs
        success_rate = (success_logs / total_logs * 100) if total_logs > 0 else 0

        # Summary data
        summary_data = [
            ("Total Log Entries", total_logs),
            ("Successful Operations", success_logs),
            ("Failed Operations", error_logs),
            ("Success Rate", f"{success_rate:.1f}%"),
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        row = 3
        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].border = self.border
            ws[f"B{row}"].border = self.border
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
